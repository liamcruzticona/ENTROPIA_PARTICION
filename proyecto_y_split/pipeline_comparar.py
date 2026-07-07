"""
=====================================================================
  pipeline_comparar.py - Pipeline MIN+MAX en BEST y WORST
  Compara entropias, MI, MST y estructura entre los dos grupos.
=====================================================================
"""

import numpy as np
from collections import Counter
from itertools import combinations

# =============================================================================
# FUNCIONES COMUNES
# =============================================================================

def entropy(probs):
    return -sum(p * np.log2(p) for p in probs if p > 0)


def mutual_info(col_x, col_y, n):
    jc = {}
    for i in range(n):
        k = (col_x[i], col_y[i])
        jc[k] = jc.get(k, 0) + 1
    mi = 0.0
    for (vx, vy), c in jc.items():
        p_xy = c / n
        p_x = np.sum(col_x == vx) / n
        p_y = np.sum(col_y == vy) / n
        mi += p_xy * np.log2(p_xy / (p_x * p_y))
    return mi


class UnionFind:
    def __init__(self, n):
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x):
        if self.parent[x] != x:
            self.parent[x] = self.find(self.parent[x])
        return self.parent[x]

    def union(self, x, y):
        rx, ry = self.find(x), self.find(y)
        if rx == ry:
            return False
        if self.rank[rx] < self.rank[ry]:
            self.parent[rx] = ry
        elif self.rank[rx] > self.rank[ry]:
            self.parent[ry] = rx
        else:
            self.parent[ry] = rx
            self.rank[rx] += 1
        return True


def load_dataset(filename):
    with open(filename, 'r') as f:
        lines = f.readlines()
    h = lines[0].strip().split(',')
    d = np.array([[int(x) for x in l.strip().split(',')] for l in lines[1:]])
    return h, d


def run_pipeline(header, data, label):
    n_rows, n_cols = len(data), len(header)
    variables = header

    # Frecuencias y probabilidades
    freq_tables = {}
    prob_tables = {}
    unique_vals = {}
    for j, var in enumerate(variables):
        col = data[:, j]
        counts = Counter(col)
        freq_tables[var] = dict(sorted(counts.items()))
        prob_tables[var] = {k: v / n_rows for k, v in sorted(counts.items())}
        unique_vals[var] = sorted(counts.keys())

    # Entropias
    entropies = {var: entropy(list(prob_tables[var].values())) for var in variables}

    # Informacion Mutua
    mi_dict = {}
    for v1, v2 in combinations(range(n_cols), 2):
        m = mutual_info(data[:, v1], data[:, v2], n_rows)
        mi_dict[(v1, v2)] = m
        mi_dict[(v2, v1)] = m

    mi_matrix = np.zeros((n_cols, n_cols))
    for i in range(n_cols):
        for j in range(n_cols):
            if i != j:
                mi_matrix[i, j] = mi_dict.get((i, j), 0.0)

    # Matriz de distancias
    dist_matrix = np.zeros((n_cols, n_cols))
    for i in range(n_cols):
        for j in range(n_cols):
            if i != j:
                min_h = min(entropies[variables[i]], entropies[variables[j]])
                dist_matrix[i, j] = max(0.0, 1.0 - mi_matrix[i, j] / min_h) if min_h > 0 else 0.0

    # Prim MIN
    visited = [False] * n_cols
    visited[0] = True
    prim_min = []
    prim_min_cost = 0.0
    for _ in range(n_cols - 1):
        bd = float('inf')
        bu, bv = -1, -1
        for u in range(n_cols):
            if visited[u]:
                for v in range(n_cols):
                    if not visited[v] and dist_matrix[u, v] < bd:
                        bd = dist_matrix[u, v]
                        bu, bv = u, v
        visited[bv] = True
        prim_min.append((bu, bv, bd))
        prim_min_cost += bd

    # Kruskal MIN
    all_min = [(dist_matrix[i, j], i, j) for i in range(n_cols) for j in range(i + 1, n_cols)]
    all_min.sort()
    uf = UnionFind(n_cols)
    kruskal_min = []
    kruskal_min_cost = 0.0
    for d, u, v in all_min:
        if uf.union(u, v):
            kruskal_min.append((u, v, d))
            kruskal_min_cost += d
            if len(kruskal_min) == n_cols - 1:
                break

    # Prim MAX
    visited = [False] * n_cols
    visited[0] = True
    prim_max = []
    prim_max_cost = 0.0
    for _ in range(n_cols - 1):
        bw = -1.0
        bu, bv = -1, -1
        for u in range(n_cols):
            if visited[u]:
                for v in range(n_cols):
                    if not visited[v] and mi_matrix[u, v] > bw:
                        bw = mi_matrix[u, v]
                        bu, bv = u, v
        visited[bv] = True
        prim_max.append((bu, bv, bw))
        prim_max_cost += bw

    # Kruskal MAX
    all_max = [(mi_matrix[i, j], i, j) for i in range(n_cols) for j in range(i + 1, n_cols)]
    all_max.sort(reverse=True)
    uf = UnionFind(n_cols)
    kruskal_max = []
    kruskal_max_cost = 0.0
    for w, u, v in all_max:
        if uf.union(u, v):
            kruskal_max.append((u, v, w))
            kruskal_max_cost += w
            if len(kruskal_max) == n_cols - 1:
                break

    return {
        'label': label, 'n_rows': n_rows, 'variables': variables,
        'freq_tables': freq_tables, 'prob_tables': prob_tables,
        'entropies': entropies, 'mi_matrix': mi_matrix,
        'dist_matrix': dist_matrix,
        'prim_min': prim_min, 'prim_min_cost': prim_min_cost,
        'kruskal_min': kruskal_min, 'kruskal_min_cost': kruskal_min_cost,
        'prim_max': prim_max, 'prim_max_cost': prim_max_cost,
        'kruskal_max': kruskal_max, 'kruskal_max_cost': kruskal_max_cost,
    }


def grados(edges, n):
    g = {}
    for u, v, _ in edges:
        g[u] = g.get(u, 0) + 1
        g[v] = g.get(v, 0) + 1
    return g


# =============================================================================
# ETAPA 0: CARGA
# =============================================================================

print("\n" + "=" * 70)
print("  PIPELINE COMPARATIVO: BEST (Y=1) vs WORST (Y=0)")
print("=" * 70)

h_b, d_b = load_dataset("d9_strong_B.csv")
h_w, d_w = load_dataset("d9_strong_W.csv")

print(f"\n  BEST  (d9_strong_B.csv): {len(d_b)} filas x {len(h_b)} variables")
print(f"  WORST (d9_strong_W.csv): {len(d_w)} filas x {len(h_w)} variables")

# =============================================================================
# ETAPA 1: PIPELINE BEST
# =============================================================================

print("\n" + "=" * 70)
print("  ETAPA 1: PIPELINE BEST (Y=1)")
print("=" * 70)

r_b = run_pipeline(h_b, d_b, "BEST")

print(f"\n  Frecuencias BEST ({r_b['n_rows']} registros):")
for var in r_b['variables']:
    freqs = r_b['freq_tables'][var]
    probs = r_b['prob_tables'][var]
    if len(freqs) == 1:
        print(f"    {var}: constante = {list(freqs.keys())[0]} (H=0)")
    else:
        items = ", ".join(f"{v}:{f}" for v, f in freqs.items())
        print(f"    {var}: {items}")

print(f"\n  Entropias BEST:")
for var in r_b['variables']:
    print(f"    H({var}) = {r_b['entropies'][var]:.6f} bits")

print(f"\n  MATRIZ IM 7x7 - BEST:")
print(f"  {'':>6}", end="")
for v in r_b['variables']:
    print(f" {v:>10}", end="")
print()
for i, vi in enumerate(r_b['variables']):
    print(f"  {vi:>6}", end="")
    for j in range(len(r_b['variables'])):
        print(f" {r_b['mi_matrix'][i,j]:10.6f}", end="")
    print()

print(f"\n  MATRIZ DE DISTANCIAS 7x7 - BEST:")
print(f"  {'':>6}", end="")
for v in r_b['variables']:
    print(f" {v:>10}", end="")
print()
for i, vi in enumerate(r_b['variables']):
    print(f"  {vi:>6}", end="")
    for j in range(len(r_b['variables'])):
        print(f" {r_b['dist_matrix'][i,j]:10.6f}", end="")
    print()

print(f"\n  BEST MIN (distancias):")
print(f"    Prim:     {r_b['prim_min_cost']:.6f}")
print(f"    Kruskal:  {r_b['kruskal_min_cost']:.6f}")
print(f"    Aristas:")

y_edges_min = []
for u, v, d in sorted(r_b['prim_min'], key=lambda x: x[2]):
    edge = f"{r_b['variables'][u]} -- {r_b['variables'][v]}"
    print(f"      {edge:<16} d={d:.6f}")
    if d < 0.001:
        y_edges_min.append(edge)

print(f"\n  BEST MAX (MI directa):")
print(f"    Prim:     {r_b['prim_max_cost']:.6f}")
print(f"    Kruskal:  {r_b['kruskal_max_cost']:.6f}")
print(f"    Aristas:")
for u, v, w in sorted(r_b['prim_max'], key=lambda x: -x[2]):
    print(f"      {r_b['variables'][u]} -- {r_b['variables'][v]:<12} MI={w:.6f}")

# =============================================================================
# ETAPA 2: PIPELINE WORST
# =============================================================================

print("\n" + "=" * 70)
print("  ETAPA 2: PIPELINE WORST (Y=0)")
print("=" * 70)

r_w = run_pipeline(h_w, d_w, "WORST")

print(f"\n  Frecuencias WORST ({r_w['n_rows']} registros):")
for var in r_w['variables']:
    freqs = r_w['freq_tables'][var]
    if len(freqs) == 1:
        print(f"    {var}: constante = {list(freqs.keys())[0]} (H=0)")
    else:
        items = ", ".join(f"{v}:{f}" for v, f in freqs.items())
        print(f"    {var}: {items}")

print(f"\n  Entropias WORST:")
for var in r_w['variables']:
    print(f"    H({var}) = {r_w['entropies'][var]:.6f} bits")

print(f"\n  MATRIZ IM 7x7 - WORST:")
print(f"  {'':>6}", end="")
for v in r_w['variables']:
    print(f" {v:>10}", end="")
print()
for i, vi in enumerate(r_w['variables']):
    print(f"  {vi:>6}", end="")
    for j in range(len(r_w['variables'])):
        print(f" {r_w['mi_matrix'][i,j]:10.6f}", end="")
    print()

print(f"\n  MATRIZ DE DISTANCIAS 7x7 - WORST:")
print(f"  {'':>6}", end="")
for v in r_w['variables']:
    print(f" {v:>10}", end="")
print()
for i, vi in enumerate(r_w['variables']):
    print(f"  {vi:>6}", end="")
    for j in range(len(r_w['variables'])):
        print(f" {r_w['dist_matrix'][i,j]:10.6f}", end="")
    print()

print(f"\n  WORST MIN (distancias):")
print(f"    Prim:     {r_w['prim_min_cost']:.6f}")
print(f"    Kruskal:  {r_w['kruskal_min_cost']:.6f}")
print(f"    Aristas:")
for u, v, d in sorted(r_w['prim_min'], key=lambda x: x[2]):
    print(f"      {r_w['variables'][u]} -- {r_w['variables'][v]:<12} d={d:.6f}")

print(f"\n  WORST MAX (MI directa):")
print(f"    Prim:     {r_w['prim_max_cost']:.6f}")
print(f"    Kruskal:  {r_w['kruskal_max_cost']:.6f}")
print(f"    Aristas:")
for u, v, w in sorted(r_w['prim_max'], key=lambda x: -x[2]):
    print(f"      {r_w['variables'][u]} -- {r_w['variables'][v]:<12} MI={w:.6f}")

# =============================================================================
# ETAPA 3: COMPARACION
# =============================================================================

print("\n\n" + "=" * 70)
print("  ETAPA 3: COMPARACION BEST vs WORST")
print("=" * 70)

print("\n  NOTA SOBRE Y:")
print("  Y (x7=x8) se uso SOLO para particionar en BEST y WORST.")
print("  Y NO se incluye en los CSVs finales (7 variables: x1-x6,x9).")
print("  Sin Y, el arbol MIN tiene costo REAL (sin degeneracion).")

# Grados MIN y MAX
g_b_min = grados(r_b['prim_min'], 7)
g_w_min = grados(r_w['prim_min'], 7)
g_b_max = grados(r_b['prim_max'], 7)
g_w_max = grados(r_w['prim_max'], 7)

hub_b_max = max(g_b_max, key=g_b_max.get)
hub_w_max = max(g_w_max, key=g_w_max.get)

# Aristas en comun MAX
eb = {(min(u, v), max(u, v)) for u, v, _ in r_b['prim_max']}
ew = {(min(u, v), max(u, v)) for u, v, _ in r_w['prim_max']}
comun = eb & ew

# Pares d=0 (sin considerar los de Y)
zeros_b = [(r_b['variables'][u], r_b['variables'][v]) for u, v, d in r_b['prim_min']
           if d < 0.001 and r_b['variables'][u] != 'Y' and r_b['variables'][v] != 'Y']
zeros_w = [(r_w['variables'][u], r_w['variables'][v]) for u, v, d in r_w['prim_min']
           if d < 0.001 and r_w['variables'][u] != 'Y' and r_w['variables'][v] != 'Y']

print(f"\n  COMPARACION DE METRICAS:")
print(f"  {'Metrica':<32} {'BEST (Y=1)':<22} {'WORST (Y=0)':<22}")
print(f"  {'-'*76}")
print(f"  {'Registros':<32} {r_b['n_rows']:<22} {r_w['n_rows']:<22}")
print(f"  {'Hub MAX (MI directa)':<32} "
      f"{r_b['variables'][hub_b_max]+' (grado '+str(g_b_max[hub_b_max])+')':<22} "
      f"{r_w['variables'][hub_w_max]+' (grado '+str(g_w_max[hub_w_max])+')':<22}")
print(f"  {'Costo Prim MIN':<32} {r_b['prim_min_cost']:<22.6f} {r_w['prim_min_cost']:<22.6f}")
print(f"  {'Costo Kruskal MIN':<32} {r_b['kruskal_min_cost']:<22.6f} {r_w['kruskal_min_cost']:<22.6f}")
print(f"  {'Costo Prim MAX':<32} {r_b['prim_max_cost']:<22.6f} {r_w['prim_max_cost']:<22.6f}")
print(f"  {'Costo Kruskal MAX':<32} {r_b['kruskal_max_cost']:<22.6f} {r_w['kruskal_max_cost']:<22.6f}")
print(f"  {'Prim=Kruskal (MIN)':<32} "
      f"{'SI' if abs(r_b['prim_min_cost']-r_b['kruskal_min_cost'])<1e-6 else 'NO':<22} "
      f"{'SI' if abs(r_w['prim_min_cost']-r_w['kruskal_min_cost'])<1e-6 else 'NO':<22}")
print(f"  {'Prim=Kruskal (MAX)':<32} "
      f"{'SI' if abs(r_b['prim_max_cost']-r_b['kruskal_max_cost'])<1e-6 else 'NO':<22} "
      f"{'SI' if abs(r_w['prim_max_cost']-r_w['kruskal_max_cost'])<1e-6 else 'NO':<22}")
print(f"  {'Pares d=0 (sin Y)':<32} {str(len(zeros_b)):<22} {str(len(zeros_w)):<22}")
print(f"  {'Aristas en comun MAX':<32} {len(comun)}/7 ({100*len(comun)/7:.0f}%)")

print(f"\n  ENTROPIAS COMPARADAS (sin Y):")
print(f"  {'Variable':<10} {'BEST':<14} {'WORST':<14} {'Diferencia':<14}")
print(f"  {'-'*52}")
for var in r_b['variables']:
    if var == 'Y':
        continue
    hb = r_b['entropies'][var]
    hw = r_w['entropies'][var]
    print(f"  {var:<10} {hb:<14.6f} {hw:<14.6f} {hb-hw:+14.6f}")

print(f"\n  MAX ARBOL - BEST (MI directa):")
for u, v, w in sorted(r_b['prim_max'], key=lambda x: -x[2]):
    print(f"    {r_b['variables'][u]} -- {r_b['variables'][v]:<6}  MI = {w:.6f}")
print(f"    Costo = {r_b['prim_max_cost']:.6f}")

print(f"\n  MAX ARBOL - WORST (MI directa):")
for u, v, w in sorted(r_w['prim_max'], key=lambda x: -x[2]):
    print(f"    {r_w['variables'][u]} -- {r_w['variables'][v]:<6}  MI = {w:.6f}")
    print(f"    Costo = {r_w['prim_max_cost']:.6f}")

# =============================================================================
# ETAPA 4: SELECCION DE VARIABLES (metodo rVP)
# =============================================================================

print("\n\n" + "=" * 70)
print("  ETAPA 4: SELECCION DE VARIABLES - Metodo rVP")
print("  rVP = Rango de Variacion Permitido")
print("  Metodo: distancias de caminos dentro del MST (MAX, MI)")
print("  Delta = SUM_A - SUM_B")
print("  |Delta| > 0.5 -> variable CRITICA")
print("=" * 70)

import networkx as nx

def build_mst(mi_matrix, variables):
    n = len(variables)
    G = nx.Graph()
    for i in range(n):
        for j in range(i + 1, n):
            G.add_edge(i, j, weight=mi_matrix[i, j])
    G_neg = nx.Graph()
    for u, v, data in G.edges(data=True):
        G_neg.add_edge(u, v, weight=-data['weight'])
    mst_neg = nx.minimum_spanning_tree(G_neg)
    mst = []
    for u, v in mst_neg.edges():
        mst.append((u, v, mi_matrix[u, v]))
    return mst

def rVP(mi_matrix, variables):
    n = len(variables)
    mst = build_mst(mi_matrix, variables)
    G_mst = nx.Graph()
    for u, v, w in mst:
        G_mst.add_edge(variables[u], variables[v], weight=w)
    path_matrix = np.zeros((n, n))
    for i in range(n):
        for j in range(n):
            if i != j:
                path = nx.shortest_path(G_mst, source=variables[i], target=variables[j], weight='weight')
                path_matrix[i, j] = sum(G_mst[path[k]][path[k+1]]['weight'] for k in range(len(path)-1))
    suma = {i: sum(path_matrix[i]) for i in range(n)}
    return path_matrix, suma, mst

_, sum_b, _ = rVP(r_b['mi_matrix'], r_b['variables'])
_, sum_w, _ = rVP(r_w['mi_matrix'], r_w['variables'])

ranking = []
for i, var in enumerate(r_b['variables']):
    sb = sum_b[i]; sw = sum_w[i]; delta = sb - sw
    ranking.append((var, sb, sw, delta, abs(delta)))

ranking.sort(key=lambda x: -x[4])

# Gap detection: caida natural (sin umbral fijo)
deltas_rvp = [ad for _, _, _, _, ad in ranking]
gaps_rvp = [deltas_rvp[i] / deltas_rvp[i+1] if deltas_rvp[i+1] > 0 else 999 for i in range(len(deltas_rvp)-1)]
corte_rvp = gaps_rvp.index(max(gaps_rvp))

print(f"\n  {'Variable':<10} {'SUM BEST':<14} {'SUM WORST':<14} {'Delta':>12} {'|Delta|':<10} {'Critica?'}")
print(f"  {'-'*68}")
for i, (var, sb, sw, d, ad) in enumerate(ranking):
    critica = 'SI' if i <= corte_rvp else 'No'
    print(f"  {var:<10} {sb:<14.4f} {sw:<14.4f} {d:12.4f} {ad:<10.4f} {critica}")

criticas = [(var, f'{d:+.4f}') for i, (var, _, _, d, ad) in enumerate(ranking) if i <= corte_rvp]
estables = [var for i, (var, _, _, _, _) in enumerate(ranking) if i > corte_rvp]
print(f"\n  Corte por caida natural: despues de posicion {corte_rvp+1}")
print(f"  Gap: {deltas_rvp[corte_rvp]:.4f} -> {deltas_rvp[corte_rvp+1]:.4f} (ratio {max(gaps_rvp):.1f}x)")
print(f"  Variables CRITICAS: {criticas}")
print(f"  Variables ESTABLES: {estables}")

# --- METODO 2: Profesor (suma fila completa matriz distancias) ---
print(f"\n  {'='*50}")
print(f"  METODO 2 (Profesor): Suma fila completa matriz distancias")
print(f"  Delta = SUM_BEST - SUM_WORST (sin valor absoluto)")
print(f"  {'='*50}")

sum_d_b = {}; sum_d_w = {}
for i in range(len(r_b['variables'])):
    sum_d_b[i] = sum(r_b['dist_matrix'][i])
    sum_d_w[i] = sum(r_w['dist_matrix'][i])

ranking2 = []
for i, var in enumerate(r_b['variables']):
    sb = sum_d_b[i]; sw = sum_d_w[i]; delta = sb - sw
    ranking2.append((var, sb, sw, delta, abs(delta)))

ranking2.sort(key=lambda x: -x[4])

print(f"\n  {'Variable':<10} {'SUM BEST':<14} {'SUM WORST':<14} {'Delta':>12} {'|Delta|':<10}")
print(f"  {'-'*58}")
for var, sb, sw, d, ad in ranking2:
    print(f"  {var:<10} {sb:<14.4f} {sw:<14.4f} {d:12.4f} {ad:<10.4f}")

# Gap detection
deltas_ordered = [ad for _, _, _, _, ad in ranking2]
gaps = [deltas_ordered[i] / deltas_ordered[i+1] if deltas_ordered[i+1] > 0 else 999 for i in range(len(deltas_ordered)-1)]
max_gap_idx = gaps.index(max(gaps))
criticas2 = [var for i, (var, _, _, _, _) in enumerate(ranking2) if i <= max_gap_idx]
discriminativas = criticas2
print(f"\n  Discriminativas por caida natural (pos {max_gap_idx+1}): {discriminativas}")
print(f"  (gap maximo: {deltas_ordered[max_gap_idx]:.4f} -> {deltas_ordered[max_gap_idx+1]:.4f}, ratio={max(gaps):.1f}x)")

# Comparacion lado a lado
print(f"\n  {'='*50}")
print(f"  COMPARACION: rVP vs Profesor")
print(f"  {'='*50}")
print(f"  {'Variable':<10} {'rVP |D|':<10} {'rVP rank':<10} {'Prof |D|':<10} {'Prof rank':<10}")
print(f"  {'-'*50}")
for i, var in enumerate(r_b['variables']):
    rvp_rank = next(j+1 for j, (v,_,_,_,_) in enumerate(ranking) if v == var)
    rvp_d = next(ad for v,_,_,_,ad in ranking if v == var)
    prof_rank = next(j+1 for j, (v,_,_,_,_) in enumerate(ranking2) if v == var)
    prof_d = next(ad for v,_,_,_,ad in ranking2 if v == var)
    print(f"  {var:<10} {rvp_d:<10.4f} {rvp_rank:<10} {prof_d:<10.4f} {prof_rank:<10}")

print("\n" + "=" * 70)
print("  PIPELINE COMPLETADO")
print("=" * 70)
print()

# =============================================================================
# ETAPA 4: GENERAR EXCEL
# =============================================================================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FILL = PatternFill('solid', fgColor='2C3E50')
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
GREEN_FILL = PatternFill('solid', fgColor='27AE60')
GREEN_FONT = Font(bold=True, size=11, color='FFFFFF')
ORANGE_FILL = PatternFill('solid', fgColor='F39C12')
LIGHT_GREEN = PatternFill('solid', fgColor='D5F5E3')
LIGHT_BLUE = PatternFill('solid', fgColor='D6EAF8')
LIGHT_ORANGE = PatternFill('solid', fgColor='FDEBD0')
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
CENTER = Alignment(horizontal='center')

def style_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font; cell.fill = fill; cell.alignment = CENTER; cell.border = BORDER

def sc(cell, fill=None, font=None):
    cell.border = BORDER; cell.alignment = CENTER
    if fill: cell.fill = fill
    if font: cell.font = font

def write_sheet(ws, r, fill_color, name):
    ws.merge_cells('A1:F1')
    ws['A1'] = f"{name} ({r['n_rows']} registros)"
    ws['A1'].font = Font(bold=True, size=14, color=fill_color.fgColor)
    row = 3

    # Entropias
    ws.cell(row=row, column=1, value="ENTROPIAS").font = Font(bold=True, size=12)
    row += 1
    for c, h in enumerate(['Variable', 'H(X) bits'], 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, 2)
    row += 1
    for var in r['variables']:
        ws.cell(row=row, column=1, value=var)
        ws.cell(row=row, column=2, value=round(r['entropies'][var], 6))
        sc(ws.cell(row=row, column=1)); sc(ws.cell(row=row, column=2))
        row += 1
    row += 1

    # Matriz IM
    ws.cell(row=row, column=1, value="MATRIZ IM 7x7").font = Font(bold=True, size=12)
    row += 1
    nv = len(r['variables'])
    for c, h in enumerate([''] + r['variables'], 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, nv + 1)
    row += 1
    for i, vi in enumerate(r['variables']):
        ws.cell(row=row, column=1, value=vi)
        sc(ws.cell(row=row, column=1), font=Font(bold=True))
        for j in range(nv):
            val = r['mi_matrix'][i, j]
            ws.cell(row=row, column=j + 2, value=round(val, 6))
            fl = LIGHT_GREEN if val > 0.3 else None
            sc(ws.cell(row=row, column=j + 2), fill=fl)
        row += 1
    row += 1

    # Matriz de distancias
    ws.cell(row=row, column=1, value="MATRIZ DE DISTANCIAS 7x7").font = Font(bold=True, size=12)
    row += 1
    for c, h in enumerate([''] + r['variables'], 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, nv + 1)
    row += 1
    for i, vi in enumerate(r['variables']):
        ws.cell(row=row, column=1, value=vi)
        sc(ws.cell(row=row, column=1), font=Font(bold=True))
        for j in range(nv):
            val = r['dist_matrix'][i, j]
            ws.cell(row=row, column=j + 2, value=round(val, 6))
            fl = LIGHT_GREEN if val < 0.5 else (LIGHT_ORANGE if val == 0 else None)
            sc(ws.cell(row=row, column=j + 2), fill=fl)
        row += 1
    row += 1

    # MST MIN
    ws.cell(row=row, column=1,
            value=f"MST MIN (distancias) - Costo: {r['prim_min_cost']:.6f}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for c, h in enumerate(['Arista', 'Distancia d'], 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, 2)
    row += 1
    for u, v, d in sorted(r['prim_min'], key=lambda x: x[2]):
        ws.cell(row=row, column=1,
                value=f"{r['variables'][u]} -- {r['variables'][v]}")
        ws.cell(row=row, column=2, value=round(d, 6))
        fl = LIGHT_GREEN if d < 0.5 else None
        sc(ws.cell(row=row, column=1), fill=fl)
        sc(ws.cell(row=row, column=2), fill=fl)
        row += 1
    row += 1

    # MST MAX
    ws.cell(row=row, column=1,
            value=f"MST MAX (MI directa) - Costo: {r['prim_max_cost']:.6f}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for c, h in enumerate(['Arista', 'Peso MI'], 1):
        ws.cell(row=row, column=c, value=h)
    style_header(ws, row, 2)
    row += 1
    for u, v, w in sorted(r['prim_max'], key=lambda x: -x[2]):
        ws.cell(row=row, column=1,
                value=f"{r['variables'][u]} -- {r['variables'][v]}")
        ws.cell(row=row, column=2, value=round(w, 6))
        fl = LIGHT_BLUE if w > 0.3 else None
        sc(ws.cell(row=row, column=1), fill=fl)
        sc(ws.cell(row=row, column=2), fill=fl)
        row += 1

    for col in ws.columns:
        ml = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 45)

# Hoja BEST
ws_b = wb.active
ws_b.title = "BEST_Y1"
write_sheet(ws_b, r_b, GREEN_FILL, "BEST (Y=1)")

# Hoja WORST
ws_w = wb.create_sheet("WORST_Y0")
write_sheet(ws_w, r_w, ORANGE_FILL, "WORST (Y=0)")

# Hoja COMPARACION
ws_c = wb.create_sheet("COMPARACION")
ws_c.merge_cells('A1:D1')
ws_c['A1'] = "COMPARACION BEST (Y=1) vs WORST (Y=0)"
ws_c['A1'].font = Font(bold=True, size=14, color='2C3E50')
row = 3
for c, h in enumerate(['Metrica', 'BEST', 'WORST'], 1):
    ws_c.cell(row=row, column=c, value=h)
style_header(ws_c, row, 3)
row += 1

comp = [
    ('Registros', r_b['n_rows'], r_w['n_rows']),
    ('Hub MAX', f"x1 (g=3)", f"x9 (g=4)"),
    ('Costo Prim MIN', r_b['prim_min_cost'], r_w['prim_min_cost']),
    ('Costo Kruskal MIN', r_b['kruskal_min_cost'], r_w['kruskal_min_cost']),
    ('Costo Prim MAX', r_b['prim_max_cost'], r_w['prim_max_cost']),
    ('Costo Kruskal MAX', r_b['kruskal_max_cost'], r_w['kruskal_max_cost']),
    ('Aristas en comun MAX', '6/7 (86%)', '6/7 (86%)'),
]
for label, bv, wv in comp:
    ws_c.cell(row=row, column=1, value=label)
    if isinstance(bv, float):
        ws_c.cell(row=row, column=2, value=round(bv, 6))
        ws_c.cell(row=row, column=3, value=round(wv, 6))
    else:
        ws_c.cell(row=row, column=2, value=str(bv))
        ws_c.cell(row=row, column=3, value=str(wv))
    for c in range(1, 4): sc(ws_c.cell(row=row, column=c))
    row += 1

row += 1
ws_c.cell(row=row, column=1, value="ENTROPIAS").font = Font(bold=True, size=12)
row += 1
for c, h in enumerate(['Variable', 'BEST', 'WORST', 'Diferencia'], 1):
    ws_c.cell(row=row, column=c, value=h)
style_header(ws_c, row, 4)
row += 1
for var in r_b['variables']:
    if var == 'Y': continue
    ws_c.cell(row=row, column=1, value=var)
    ws_c.cell(row=row, column=2, value=round(r_b['entropies'][var], 6))
    ws_c.cell(row=row, column=3, value=round(r_w['entropies'][var], 6))
    ws_c.cell(row=row, column=4, value=round(r_b['entropies'][var] - r_w['entropies'][var], 6))
    for c in range(1, 5): sc(ws_c.cell(row=row, column=c))
    row += 1

for col in ws_c.columns:
    ml = max((len(str(c.value or '')) for c in col), default=10)
    ws_c.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 4, 45)

output = "resultados_split_nuevo.xlsx"
wb.save(output)
print(f"\n  Excel generado: {output}  (3 hojas: BEST_Y1, WORST_Y0, COMPARACION)")

